#!/usr/bin/env python3
"""gen-uat-xlsx.py - 人手E2E（UAT）シナリオを Excel ワークブックに一方向生成する。

手順のSSOT（Single Source of Truth）は story{N}.md（Markdown）であり、
本スクリプトはそこから xlsx を一方向生成する。実施結果（OK/NG・証跡）は xlsx に記入する。

アンカー仕様:
    story{N}.md 内の `## SC-...` 見出しを起点とするSCセクションのうち、
    本文に `**種別**: 人手E2E` を含むものをパース対象とする。
    対象SC直後の「### テスト手順」テーブル（`| No | 手順 | 確認項目 | 期待値 | 結果 | 備考 |`）を読み取る。
    区切り行（`|---|...`）はスキップする。

列マッピング（Markdown 6列 → xlsx 7列）:
    Markdown: No / 手順 / 確認項目 / 期待値 / 結果 / 備考
    xlsx:     No / 手順 / 確認項目 / 期待値 / 結果[OK/NG/保留 ドロップダウン] / 証跡ファイル名[新規・空] / 備考
    （Markdownの「結果」列の値は転記せず、xlsxではドロップダウン入力欄にする。
      「証跡ファイル名」列はMarkdownに対応列がないため新規・空欄で追加する。）

参考: Anthropic公式 xlsx スキル（github.com/anthropics/skills の skills/xlsx, openpyxl使用）の
アプローチを参考にしたが、コードは再配布不可のため流用せず独自実装している。

Usage:
    gen-uat-xlsx.py <spec_dir> [--out <path>] [--issue <N>]
"""

import argparse
import glob
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print(
        "openpyxl が必要です。`pip install openpyxl` を実行してください。"
        "手順自体は story{N}.md（Markdown）でそのまま参照可能です。",
        file=sys.stderr,
    )
    sys.exit(1)


# ---- 定数 ----------------------------------------------------------------

RESULT_CHOICES = '"OK,NG,保留"'
XLSX_HEADERS = ["No", "手順", "確認項目", "期待値", "結果", "証跡ファイル名", "備考"]
# story1.md など複数 story を対象にするための glob パターン
STORY_GLOB = "story*.md"

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
META_FONT = Font(bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


# ---- パース ---------------------------------------------------------------


def derive_issue_number(spec_dir: str, override: "str | None") -> str:
    """spec_dir のパスから issue 番号 N を抽出する。取れなければ basename を使う。"""
    if override:
        return override
    m = re.search(r"issue(\d+)", spec_dir)
    if m:
        return m.group(1)
    base = os.path.basename(os.path.normpath(spec_dir))
    return base or "unknown"


def split_sc_sections(text: str):
    """Markdown を `## SC-...` 見出し単位のセクションに分割する。

    返り値: [(sc_title, body_text), ...]
    """
    sections = []
    current_title = None
    current_lines: list = []
    for line in text.splitlines():
        m = re.match(r"^##\s+(SC-\S.*)$", line)
        if m:
            if current_title is not None:
                sections.append((current_title, "\n".join(current_lines)))
            current_title = m.group(1).strip()
            current_lines = []
        elif current_title is not None:
            current_lines.append(line)
    if current_title is not None:
        sections.append((current_title, "\n".join(current_lines)))
    return sections


def is_manual_e2e(body: str) -> bool:
    """SC本文が人手E2E種別かどうか（`**種別**: 人手E2E` を含むか）を判定する。"""
    return bool(re.search(r"\*\*種別\*\*\s*[:：]\s*人手E2E", body))


def parse_test_steps(body: str):
    """SC本文から「### テスト手順」テーブルの行をパースする。

    返り値: [{"No":..,"手順":..,"確認項目":..,"期待値":..,"備考":..}, ...]
    （結果列の値は転記しないため取り込まない。）
    """
    lines = body.splitlines()
    # 「### テスト手順」見出しを探す
    start = None
    for i, line in enumerate(lines):
        if re.match(r"^###\s+テスト手順", line):
            start = i + 1
            break
    if start is None:
        return []

    rows = []
    seen_header = False
    for line in lines[start:]:
        stripped = line.strip()
        if not stripped.startswith("|"):
            # ヘッダ検出後に非パイプ行（空行・次の見出し等）が来たらテーブル終端とみなす。
            # 空行を読み飛ばすと後続の無関係なパイプ表を誤取り込みするため break する。
            if seen_header:
                break
            continue
        # 区切り行 (|---|---|) をスキップ
        if re.match(r"^\|[\s:|-]+\|?\s*$", stripped):
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        # ヘッダ行（No 手順 ...）をスキップ
        if not seen_header and cells and cells[0] in ("No", "no", "NO"):
            seen_header = True
            continue
        if not seen_header:
            # ヘッダ未検出のまま本文行が来た場合も一応ヘッダ扱いでスキップしない
            seen_header = True
        # 列数チェック: 最低 No・手順 の2列を満たさない行はスキップ（表崩れ検知）
        if len(cells) < 2:
            print(
                f"  [警告] 列数不足の手順行をスキップしました: {stripped}",
                file=sys.stderr,
            )
            continue
        # 6列想定: No 手順 確認項目 期待値 結果 備考
        def col(idx):
            return cells[idx] if idx < len(cells) else ""

        rows.append(
            {
                "No": col(0),
                "手順": col(1),
                "確認項目": col(2),
                "期待値": col(3),
                "備考": col(5),
            }
        )
    return rows


def collect_manual_e2e_scenarios(spec_dir: str):
    """spec_dir/qa-tests/story*.md から人手E2Eシナリオを収集する。

    返り値: [{"story_file":.., "sc_title":.., "steps":[...]}, ...]
    """
    qa_dir = os.path.join(spec_dir, "qa-tests")
    pattern = os.path.join(qa_dir, STORY_GLOB)
    story_files = sorted(glob.glob(pattern))
    scenarios = []
    for sf in story_files:
        with open(sf, encoding="utf-8") as fp:
            text = fp.read()
        for sc_title, body in split_sc_sections(text):
            if not is_manual_e2e(body):
                continue
            steps = parse_test_steps(body)
            scenarios.append(
                {
                    "story_file": os.path.basename(sf),
                    "sc_title": sc_title,
                    "steps": steps,
                }
            )
    return scenarios


# ---- xlsx 生成 ------------------------------------------------------------


def sanitize_sheet_name(name: str) -> str:
    """Excel シート名に使えない文字を除去し、31文字に丸める。"""
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name)
    cleaned = cleaned.strip() or "sheet"
    return cleaned[:31]


def write_legend_sheet(ws):
    ws.title = "凡例・記入方法"
    ws.column_dimensions["A"].width = 100
    rows = [
        "■ このワークブックについて",
        "マージ/デプロイ後に人間が実施する『人手E2E』シナリオの打鍵記録用です。",
        "手順の正本（SSOT）は story{N}.md（Markdown）です。本ファイルは手順を一方向で転記したものです。",
        "手順そのものを変更したい場合は story{N}.md を修正し、本ファイルを再生成してください。",
        "",
        "■ 実施手順",
        "1. 各シナリオの『手順マスタ』タブを右クリック → 『移動またはコピー』→ 『コピーを作成』で複製します。",
        "2. 複製したタブを『テスト実施_{名前}_{YYYYMMDD}』にリネームします（例: テスト実施_山田_20260617）。",
        "3. ヘッダ部の 実施日 / 実施者 / 環境URL を記入します。",
        "4. 各手順の『結果』列でOK/NG/保留を選びます（ドロップダウン）。",
        "5. 証跡（スクショ等）を撮ったら『証跡ファイル名』列にファイル名を記入します。",
        "6. 最後に『全体ステータス』を記入します。",
        "",
        "■ ステータス定義",
        "OK   : 期待値どおりの結果が得られた",
        "NG   : 期待値と異なる結果（不具合・要修正）",
        "保留 : 環境都合等で実施できない / 判断保留",
        "",
        "■ 証跡の保存先と命名",
        "保存先: qa-tests/evidence/ 配下に保存してください。",
        "命名例: SC-06_no2_NG_20260617.png のように シナリオID_手順No_結果_日付 を推奨します。",
        "『証跡ファイル名』列には保存したファイル名を記入します。",
        "",
        "■ 注意",
        "・『手順マスタ』タブは雛形です。直接記入せず必ず複製してから記入してください。",
        "・1人1タブ（または1実施1タブ）を基本とし、タブ名の重複に注意してください。",
    ]
    for i, line in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("■"):
            cell.font = Font(bold=True, size=12)


def write_master_sheet(wb, scenario, used_names):
    sc_title = scenario["sc_title"]
    # SC-06 のような ID を抽出してシート名に使う
    m = re.match(r"(SC-\d+)", sc_title)
    sc_id = m.group(1) if m else sc_title
    base_name = sanitize_sheet_name(f"手順マスタ_{sc_id}")
    name = base_name
    n = 2
    while name in used_names:
        suffix = f"_{n}"
        name = sanitize_sheet_name(base_name[: 31 - len(suffix)] + suffix)
        n += 1
    used_names.add(name)

    ws = wb.create_sheet(title=name)

    # ヘッダ部（記入欄）
    ws["A1"] = "シナリオ"
    ws["A1"].font = META_FONT
    ws["B1"] = sc_title
    meta = [
        ("実施日", ""),
        ("実施者", ""),
        ("環境URL", ""),
        ("全体ステータス", ""),
    ]
    for i, (label, val) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = META_FONT
        ws.cell(row=i, column=2, value=val)

    header_row = len(meta) + 3  # 1行あけてからヘッダ

    # 表ヘッダ（7列）
    for c, h in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    steps = scenario["steps"]
    first_data_row = header_row + 1
    for r, step in enumerate(steps, start=first_data_row):
        ws.cell(row=r, column=1, value=step["No"]).border = BORDER
        ws.cell(row=r, column=2, value=step["手順"]).border = BORDER
        ws.cell(row=r, column=3, value=step["確認項目"]).border = BORDER
        ws.cell(row=r, column=4, value=step["期待値"]).border = BORDER
        ws.cell(row=r, column=5, value="").border = BORDER  # 結果（ドロップダウン）
        ws.cell(row=r, column=6, value="").border = BORDER  # 証跡ファイル名（空）
        ws.cell(row=r, column=7, value=step["備考"]).border = BORDER
        for c in range(1, 8):
            ws.cell(row=r, column=c).alignment = WRAP_TOP

    last_data_row = first_data_row + len(steps) - 1

    # 結果列(E)のドロップダウン
    if steps:
        dv = DataValidation(type="list", formula1=RESULT_CHOICES, allow_blank=True)
        dv.error = "OK / NG / 保留 のいずれかを選択してください"
        dv.errorTitle = "入力エラー"
        dv.prompt = "結果を選択してください"
        dv.promptTitle = "結果"
        ws.add_data_validation(dv)
        dv.add(f"E{first_data_row}:E{last_data_row}")

    # 列幅
    widths = {"A": 6, "B": 40, "C": 24, "D": 36, "E": 12, "F": 24, "G": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return name, len(steps)


def build_workbook(scenarios, out_path, issue_no):
    wb = openpyxl.Workbook()
    legend = wb.active
    write_legend_sheet(legend)

    used_names = {legend.title}
    summary = []
    for sc in scenarios:
        name, count = write_master_sheet(wb, sc, used_names)
        summary.append((sc["sc_title"], name, count, sc["story_file"]))

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return summary


# ---- main -----------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="人手E2E（UAT）シナリオを story{N}.md から Excel ワークブックに一方向生成する"
    )
    parser.add_argument("spec_dir", help="仕様ディレクトリ（例: docs/specs/issues/issue42-xxx/）")
    parser.add_argument("--out", help="出力xlsxパス（省略時は <spec_dir>/qa-tests/ に生成）")
    parser.add_argument("--issue", help="Issue番号を明示指定（パスから取得できない場合）")
    args = parser.parse_args()

    spec_dir = args.spec_dir
    if not os.path.isdir(spec_dir):
        print(f"spec_dir が見つかりません: {spec_dir}", file=sys.stderr)
        sys.exit(1)

    issue_no = derive_issue_number(spec_dir, args.issue)

    try:
        scenarios = collect_manual_e2e_scenarios(spec_dir)
    except Exception as e:  # best-effort: トレースバックでなく平易なメッセージ
        print(f"story{{N}}.md のパースに失敗しました: {e}", file=sys.stderr)
        sys.exit(1)

    if not scenarios:
        # 「対象なし」は正常ケース（人手E2Eを持たないIssueも多い）。
        # CI/上位Skillから呼ばれた際に失敗誤認しないよう、警告＋正常終了(0)とする。
        print(
            "人手E2E（`**種別**: 人手E2E`）シナリオが "
            f"{os.path.join(spec_dir, 'qa-tests', STORY_GLOB)} に見つかりませんでした。"
            " 生成対象が無いためスキップします（story{N}.md に `**種別**: 人手E2E` の"
            "シナリオがある場合のみ xlsx を生成します）。",
            file=sys.stderr,
        )
        sys.exit(0)

    out_path = args.out or os.path.join(
        spec_dir, "qa-tests", f"手動シナリオテスト_Issue{issue_no}.xlsx"
    )

    try:
        summary = build_workbook(scenarios, out_path, issue_no)
    except Exception as e:  # best-effort
        print(f"xlsx の生成に失敗しました: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"生成完了: {out_path}")
    print(f"人手E2Eシナリオ {len(summary)} 件をパースしました:")
    for sc_title, sheet, count, story_file in summary:
        print(f"  - {sc_title}（{story_file}）→ シート『{sheet}』 手順 {count} 件")


if __name__ == "__main__":
    main()
